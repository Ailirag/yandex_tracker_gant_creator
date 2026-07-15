"""Отчёты планировщика: xlsx с diff, машиночитаемый план, бэкап значений."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .scheduler import PlanResult
from .shifts import Shift

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4EC")


def _d(value: date | None) -> str:
    return value.isoformat() if value else ""


def write_xlsx(
    result: PlanResult,
    out_path: Path,
    capacity_note: str,
    shifts: list[Shift] | None = None,
) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "План"
    headers = [
        "Ключ", "Название", "Команда", "Системы", "Статус", "Важность",
        "СА, ч", "СА старт", "СА конец",
        "Dev, ч", "Dev старт", "Dev конец",
        "Тест-буфер до", "Релиз (Р)", "Р из фолбэка",
        "start: было", "start: расчёт", "end: было", "end: расчёт",
        "ПДЗ: было", "ПДЗ: расчёт", "Дедлайн", "Предупреждения",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for item in result.planned:
        i = item.issue
        row = [
            i.key, i.summary, item.team.component if item.team else "",
            ", ".join(item.systems), i.status,
            i.importance if i.importance is not None else "",
            item.sa.hours if item.sa else "", _d(item.sa.start if item.sa else None),
            _d(item.sa.end if item.sa else None),
            item.dev.hours if item.dev else "", _d(item.dev.start if item.dev else None),
            _d(item.dev.end if item.dev else None),
            _d(item.buffer_end), _d(item.release_date),
            "да" if item.release_fallback else "",
            _d(i.start), _d(item.new_start), _d(i.end), _d(item.new_end),
            _d(i.planned_completion), _d(item.new_pdz), _d(i.deadline),
            "; ".join(item.warnings),
        ]
        ws.append(row)
        r = ws.max_row
        for col_was, col_new in ((16, 17), (18, 19), (20, 21)):
            if row[col_was - 1] != row[col_new - 1]:
                ws.cell(row=r, column=col_new).fill = CHANGED_FILL
        if item.warnings:
            ws.cell(row=r, column=len(headers)).fill = WARN_FILL

    ws2 = wb.create_sheet("Пропущенные")
    ws2.append(["Ключ", "Название", "Статус", "Важность", "Причина"])
    for c in range(1, 6):
        ws2.cell(row=1, column=c).font = Font(bold=True)
        ws2.cell(row=1, column=c).fill = HEADER_FILL
    for s in result.skipped:
        ws2.append([
            s.issue.key, s.issue.summary, s.issue.status,
            s.issue.importance if s.issue.importance is not None else "", s.reason,
        ])

    if shifts is not None:
        wsh = wb.create_sheet("Смещения")
        wsh.append([
            "Ключ", "Название", "Команда",
            "end: было", "end: стало", "Δ раб. дней",
            "ПДЗ: было", "ПДЗ: стало", "Вероятная причина",
        ])
        for c in range(1, 10):
            wsh.cell(row=1, column=c).font = Font(bold=True)
            wsh.cell(row=1, column=c).fill = HEADER_FILL
        for s in shifts:
            wsh.append([
                s.key, s.summary, s.team,
                _d(s.old_end), _d(s.new_end), s.delta_workdays,
                _d(s.old_pdz), _d(s.new_pdz), s.suspects,
            ])
            if s.delta_workdays > 0:
                wsh.cell(row=wsh.max_row, column=6).fill = WARN_FILL

    ws3 = wb.create_sheet("Инфо")
    ws3.append(["Сформирован", datetime.now().isoformat(timespec="seconds")])
    ws3.append(["Источник ёмкости", capacity_note])
    for w in result.warnings:
        ws3.append(["Предупреждение", w])
    for status, count in sorted(result.unrecognized_statuses.items()):
        ws3.append(["Нераспознанный статус", f"{status}: {count} задач(и)"])

    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 2, 60)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_plan_json(result: PlanResult, out_path: Path) -> None:
    """Машиночитаемый план — вход для apply и baseline для сравнения планов."""
    payload = []
    for item in result.planned:
        payload.append({
            "key": item.issue.key,
            "order": item.order,
            "team": item.team.id if item.team else None,
            "importance": item.issue.importance,
            "status": item.issue.status,
            "new_start": _d(item.new_start) or None,
            "new_end": _d(item.new_end) or None,
            "new_planned_completion": _d(item.new_pdz) or None,
            "warnings": item.warnings,
        })
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")


def issues_to_snapshot(issues) -> list[dict]:
    """Снимок текущих дат задач (для бэкапа и восстановления)."""
    return [
        {
            "key": i.key,
            "start": _d(i.start) or None,
            "end": _d(i.end) or None,
            "planned_completion": _d(i.planned_completion) or None,
            "deadline": _d(i.deadline) or None,
            "status": i.status,
        }
        for i in issues
    ]


def dump_json(payload, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")


def write_backup(result: PlanResult, out_path: Path) -> None:
    """Текущие значения дат из Трекера (по задачам плана) — снимок при расчёте."""
    dump_json(issues_to_snapshot([item.issue for item in result.planned]), out_path)


def console_summary(result: PlanResult, capacity_note: str) -> str:
    lines = [
        f"Запланировано: {len(result.planned)}, пропущено: {len(result.skipped)}, "
        f"закрыто/отменено: {result.done_count}",
        f"Источник ёмкости: {capacity_note}",
    ]
    changed = sum(
        1 for it in result.planned
        if (it.new_start, it.new_end, it.new_pdz)
        != (it.issue.start, it.issue.end, it.issue.planned_completion)
    )
    lines.append(f"Задач с изменёнными датами: {changed}")
    with_warn = sum(1 for it in result.planned if it.warnings)
    if with_warn:
        lines.append(f"Задач с предупреждениями: {with_warn}")
    for w in result.warnings:
        lines.append(f"! {w}")
    for status, count in sorted(result.unrecognized_statuses.items()):
        lines.append(f"? Нераспознанный статус «{status}»: {count}")
    return "\n".join(lines)
