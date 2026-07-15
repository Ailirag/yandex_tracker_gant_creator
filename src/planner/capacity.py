"""Источники ёмкости (FTE по дням) для групп «роль × направление».

Источники:
- FileCapacity — выгрузка из 1С (запрос onec/Запрос_ГрафикДоступности.txt):
  реальная доступность с учётом отпусков; рекомендуемый источник;
- OneCHttpCapacity — боевой: HTTP-сервис 1С отдаёт те же остатки онлайн;
- MockCapacity — до появления данных: численность групп из
  «Сотрудники в иерархии.xlsx» × коэффициент резерва × рабочие дни.

Контракт: capacity(group, day) -> FTE (float >= 0). Планировщик сам
уменьшает остаток при раскладке через CapacityLedger.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Protocol

import httpx
import openpyxl

from .calendar_ru import is_workday
from .config import RESERVE_COEFF_DEV, RESERVE_COEFF_SA, TEAMS


class CapacitySource(Protocol):
    def fte(self, group: str, day: date) -> float: ...
    def known_groups(self) -> set[str]: ...
    def describe(self) -> str: ...


# --- Мок из xlsx ---------------------------------------------------------

_SA_GROUPS = {t.sa_group for t in TEAMS if t.sa_group}
_DEV_GROUPS = {t.dev_group for t in TEAMS if t.dev_group}


@dataclass
class MockCapacity:
    """Ёмкость = активная численность группы × резервный коэффициент.

    Сотрудник учитывается, если не помечен «Исключить из ресурса» и период
    работы (ДатаНачалаРаботы/ДатаОкончанияРаботы) покрывает день. Отпуска мок
    не видит — это осознанная погрешность до подключения HTTP-сервиса 1С.
    """
    headcount: dict[str, list[tuple[date | None, date | None]]]

    @classmethod
    def from_xlsx(cls, xlsx_path) -> "MockCapacity":
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        def parse_d(v) -> date | None:
            if v is None or str(v).startswith("<"):
                return None
            if isinstance(v, date):
                return v
            try:
                dd, mm, yy = str(v).strip().split(".")
                return date(int(yy), int(mm), int(dd))
            except Exception:
                return None

        headcount: dict[str, list[tuple[date | None, date | None]]] = {}
        for row in rows:
            parent, ref, _, deleted, is_group, _, name = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
            excluded = row[13] if len(row) > 13 else None
            if ref is None or str(is_group) == "Да":
                continue
            if str(deleted) == "Да" or str(excluded) == "Да":
                continue
            start = parse_d(row[10] if len(row) > 10 else None)
            end = parse_d(row[12] if len(row) > 12 else None)
            headcount.setdefault(str(parent).strip(), []).append((start, end))
        return cls(headcount=headcount)

    def _match_group(self, group: str) -> str | None:
        """Имена групп в 1С и xlsx могут отличаться суффиксом «(1С)» и пробелами."""
        norm = lambda s: s.replace("(1С)", "").replace(" ", "").lower()
        for known in self.headcount:
            if norm(known) == norm(group):
                return known
        return None

    def fte(self, group: str, day: date) -> float:
        if not is_workday(day):
            return 0.0
        known = self._match_group(group)
        if known is None:
            return 0.0
        active = 0
        for start, end in self.headcount[known]:
            if start and day < start:
                continue
            if end and day > end:
                continue
            active += 1
        coeff = RESERVE_COEFF_SA if group in _SA_GROUPS else RESERVE_COEFF_DEV
        return active * coeff

    def known_groups(self) -> set[str]:
        return {g for g in (_SA_GROUPS | _DEV_GROUPS) if self._match_group(g)}

    def describe(self) -> str:
        return (
            "МОК: численность групп из xlsx × резерв "
            f"(Dev {RESERVE_COEFF_DEV:.0%}, СА {RESERVE_COEFF_SA:.0%}); отпуска не учтены"
        )


# --- HTTP-сервис 1С ------------------------------------------------------

@dataclass
class OneCHttpCapacity:
    """Клиент будущего HTTP-сервиса 1С (см. onec/README.md).

    GET {base_url}/capacity?from=YYYY-MM-DD&to=YYYY-MM-DD ->
    {"groups": {"<Группа>": {"YYYY-MM-DD": 3.5, ...}, ...}}
    """
    base_url: str
    date_from: date
    date_to: date
    _cache: dict[str, dict[str, float]] = field(default_factory=dict)

    def _load(self) -> None:
        if self._cache:
            return
        resp = httpx.get(
            f"{self.base_url.rstrip('/')}/capacity",
            params={"from": self.date_from.isoformat(), "to": self.date_to.isoformat()},
            timeout=120,
        )
        resp.raise_for_status()
        self._cache = resp.json()["groups"]

    def fte(self, group: str, day: date) -> float:
        self._load()
        return float(self._cache.get(group, {}).get(day.isoformat(), 0.0))

    def known_groups(self) -> set[str]:
        self._load()
        return set(self._cache)

    def describe(self) -> str:
        return f"HTTP-сервис 1С: {self.base_url} (остатки РН ДоступностьРесурса)"


# --- Выгрузка из 1С (файл) -----------------------------------------------

def _norm_group(name: str) -> str:
    return name.replace("(1С)", "").replace(" ", "").lower()


def _coerce_date(value) -> date | None:
    if value is None or isinstance(value, str) and value.startswith("<"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for parse in (
        lambda x: date.fromisoformat(x[:10]),
        lambda x: date(*(int(p) for p in reversed(x.split(".")))),  # ДД.ММ.ГГГГ
    ):
        try:
            return parse(s)
        except Exception:
            continue
    return None


@dataclass
class FileCapacity:
    """Ёмкость из выгрузки 1С (запрос onec/Запрос_ГрафикДоступности.txt).

    Ожидаемые колонки (по заголовку, регистронезависимо): Группа, Дата,
    ДоступныйРесурс. Прочие (Роль, Направление) игнорируются. Значения FTE —
    как есть из 1С (отпуска и производственный календарь уже учтены); резерв
    на поддержку/fast-track моделируется в 1С отдельными событиями, здесь
    коэффициенты НЕ применяются.
    """
    by_group_day: dict[tuple[str, str], float]
    groups: set[str]
    days: int
    date_min: date | None
    date_max: date | None
    source_name: str

    @classmethod
    def from_file(cls, path) -> "FileCapacity":
        path = Path(path)
        if path.suffix.lower() == ".json":
            rows = cls._rows_from_json(path)
        elif path.suffix.lower() in (".csv", ".tsv"):
            rows = cls._rows_from_csv(path)
        else:
            rows = cls._rows_from_xlsx(path)

        by_group_day: dict[tuple[str, str], float] = {}
        groups: set[str] = set()
        dates: set[date] = set()
        for group, day, fte in rows:
            if not group or day is None:
                continue
            by_group_day[(_norm_group(group), day.isoformat())] = float(fte or 0.0)
            groups.add(group.strip())
            dates.add(day)
        return cls(
            by_group_day=by_group_day,
            groups=groups,
            days=len(dates),
            date_min=min(dates) if dates else None,
            date_max=max(dates) if dates else None,
            source_name=path.name,
        )

    @staticmethod
    def _header_index(header: list[str]) -> dict[str, int]:
        idx: dict[str, int] = {}
        for i, cell in enumerate(header):
            key = str(cell or "").strip().lower()
            if key.startswith("групп"):
                idx["group"] = i
            elif key.startswith("дата"):
                idx["date"] = i
            elif key.startswith("доступн") or key == "fte":
                idx["fte"] = i
        missing = {"group", "date", "fte"} - set(idx)
        if missing:
            raise ValueError(
                f"В выгрузке не найдены колонки {missing}. "
                f"Ожидаются: Группа, Дата, ДоступныйРесурс. Заголовок: {header}"
            )
        return idx

    @classmethod
    def _rows_from_xlsx(cls, path: Path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # заголовок — первая строка, где есть и «Групп…», и «Дата»
        header_row = 0
        for r, row in enumerate(all_rows):
            low = [str(c or "").strip().lower() for c in row]
            if any(c.startswith("групп") for c in low) and any(c.startswith("дата") for c in low):
                header_row = r
                break
        idx = cls._header_index(list(all_rows[header_row]))
        for row in all_rows[header_row + 1:]:
            if row is None or all(c is None for c in row):
                continue
            yield (
                _cell(row, idx["group"]),
                _coerce_date(_cell(row, idx["date"])),
                _cell(row, idx["fte"]),
            )

    @classmethod
    def _rows_from_csv(cls, path: Path):
        import csv
        delim = "\t" if path.suffix.lower() == ".tsv" else ";"
        text = path.read_text(encoding="utf-8-sig")
        # автоопределение разделителя: ; или ,
        if delim == ";" and text.splitlines() and "," in text.splitlines()[0] and ";" not in text.splitlines()[0]:
            delim = ","
        reader = list(csv.reader(text.splitlines(), delimiter=delim))
        idx = cls._header_index(reader[0])
        for row in reader[1:]:
            if not row:
                continue
            yield (_cell(row, idx["group"]), _coerce_date(_cell(row, idx["date"])),
                   _fnum(_cell(row, idx["fte"])))

    @classmethod
    def _rows_from_json(cls, path: Path):
        data = json.loads(path.read_text(encoding="utf-8"))
        for item in data:
            g = item.get("Группа") or item.get("group")
            d = _coerce_date(item.get("Дата") or item.get("date"))
            f = item.get("ДоступныйРесурс") or item.get("fte") or 0
            yield (g, d, _fnum(f))

    def fte(self, group: str, day: date) -> float:
        return self.by_group_day.get((_norm_group(group), day.isoformat()), 0.0)

    def known_groups(self) -> set[str]:
        return set(self.groups)

    def describe(self) -> str:
        rng = (f"{self.date_min}..{self.date_max}" if self.date_min else "нет дат")
        return (
            f"Выгрузка 1С «{self.source_name}»: групп {len(self.groups)}, "
            f"дней {self.days} ({rng}); значения как есть (отпуска учтены в 1С)"
        )


def _cell(row, i):
    return row[i] if i < len(row) else None


def _fnum(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except ValueError:
        return 0.0


# --- Леджер остатков для раскладки ---------------------------------------

class CapacityLedger:
    """Изменяемые остатки ёмкости поверх источника (жадная раскладка)."""

    def __init__(self, source: CapacitySource):
        self._source = source
        self._used: dict[tuple[str, date], float] = {}

    def has_group(self, group: str) -> bool:
        """Есть ли у источника данные по группе (с нормализацией имени)."""
        norm = _norm_group(group)
        return any(_norm_group(g) == norm for g in self._source.known_groups())

    def available(self, group: str, day: date) -> float:
        base = self._source.fte(group, day)
        return max(0.0, base - self._used.get((group, day), 0.0))

    def consume(self, group: str, day: date, fte: float) -> None:
        self._used[(group, day)] = self._used.get((group, day), 0.0) + fte

    def utilization(self) -> dict[tuple[str, date], float]:
        return dict(self._used)
