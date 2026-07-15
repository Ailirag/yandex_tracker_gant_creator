"""Парсер «План релизов 1С.xlsx»: дни релизов и фризов по системам.

Структура листа-года: строка 1 — месяцы (значение в первой колонке месяца),
строка 2 — день месяца, далее пары строк по системам; пометки в основной
строке системы: строчные "ф","р","и","з" — дни моратория на деплой (слово
«фриз» по буквам), заглавная "Р" — день релиза. Прочие одиночные пометки
("В","О") собираются как информационные.

Факт по файлу (июль 2026): «Р» проставлены только до 29.04.2026. Для дат за
пределами заполненного плана применяется фолбэк — еженедельный релиз по
средам (доминирующий день недели в заполненной части), с предупреждением.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, timedelta

import openpyxl

RELEASE_MARK = "Р"
FREEZE_MARKS = {"ф", "р", "и", "з"}
FALLBACK_RELEASE_WEEKDAY = 2  # среда

MONTHS_RU = {
    "Январь": 1, "Февраль": 2, "Март": 3, "Апрель": 4, "Май": 5, "Июнь": 6,
    "Июль": 7, "Август": 8, "Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11,
    "Декабрь": 12,
}


@dataclass
class SystemReleases:
    """Календарь релизов одной системы."""
    name: str
    releases: list[date] = field(default_factory=list)
    freeze_days: set[date] = field(default_factory=set)
    other_marks: dict[date, str] = field(default_factory=dict)

    @property
    def plan_filled_until(self) -> date | None:
        return max(self.releases) if self.releases else None


@dataclass
class ReleasePlan:
    year: int
    systems: dict[str, SystemReleases]
    warnings: list[str] = field(default_factory=list)

    def next_release(self, system: str, not_before: date) -> tuple[date | None, bool]:
        """Ближайший день «Р» системы >= not_before.

        Возвращает (дата, использован_ли_фолбэк). Если системы нет в плане —
        (None, False): вызывающий решает, что делать.
        """
        sys_plan = self.systems.get(system)
        if sys_plan is None:
            return None, False
        for d in sys_plan.releases:
            if d >= not_before:
                return d, False
        # за пределами заполненного плана — еженедельная среда
        d = not_before
        while d.weekday() != FALLBACK_RELEASE_WEEKDAY:
            d += timedelta(days=1)
        return d, True


def parse_release_plan(xlsx_path, year: int = 2026) -> ReleasePlan:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[str(year)]

    month_starts: list[tuple[int, int]] = []  # (колонка, номер месяца)
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and str(v).strip() in MONTHS_RU:
            month_starts.append((c, MONTHS_RU[str(v).strip()]))

    def col_to_date(c: int) -> date | None:
        month = None
        for mc, m in month_starts:
            if c >= mc:
                month = m
        day = ws.cell(row=2, column=c).value
        if month is None or day is None:
            return None
        try:
            return date(year, month, int(day))
        except ValueError:
            return None

    systems: dict[str, SystemReleases] = {}
    warnings: list[str] = []

    for r in range(4, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or not str(name).strip():
            continue
        sys_rel = SystemReleases(name=str(name).strip())
        for c in range(3, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            d = col_to_date(c)
            if d is None:
                continue
            mark = str(v).strip()
            if mark == RELEASE_MARK:
                sys_rel.releases.append(d)
            elif mark in FREEZE_MARKS:
                sys_rel.freeze_days.add(d)
            else:
                sys_rel.other_marks[d] = mark
        sys_rel.releases.sort()
        if sys_rel.releases or sys_rel.freeze_days:
            systems[sys_rel.name] = sys_rel

    wb.close()

    filled = [s.plan_filled_until for s in systems.values() if s.plan_filled_until]
    if filled:
        last = max(filled)
        warnings.append(
            f"План релизов заполнен до {last.isoformat()}; для более поздних дат "
            f"используется фолбэк: релиз еженедельно по средам."
        )
    return ReleasePlan(year=year, systems=systems, warnings=warnings)
